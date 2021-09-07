from openpyxl import load_workbook
import numpy as np


def create_workbook(radius):
    file_path = R'C:\Users\manri\Documents\SAP2000_Python\INPUT\TMCLK_1Ring_TEST_raw.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Joint Coordinates']  # or wb.active
    for row in np.linspace(4,375,372,dtype=int):
        cell = 'D'+str(int(row))
        ws[cell] = radius
    file_path = R'C:\Users\manri\Documents\SAP2000_Python\INPUT\TMCLK_1Ring_TEST_raw'+str(radius)+'.xlsx'
    wb.save(file_path)

radi_options = np.linspace(1,10,10, dtype=int)
for radius in radi_options:
    create_workbook(radius)